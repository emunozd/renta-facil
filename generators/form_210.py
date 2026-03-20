"""
generators/form_210.py
Genera el borrador del Formulario 210 en Excel con:
- Valor calculado por casilla
- Explicacion de cada campo
- Destacar campos que requieren verificacion manual
"""
import logging
from typing import Optional

from interfaces.base import IFormGenerator
from config.constants import UVT

logger = logging.getLogger(__name__)

# Mapa completo de casillas con nombre y descripcion para el Excel
_CASILLAS_210 = {
    # Patrimonio
    "c29_patrimonio_bruto": {
        "numero": 29,
        "nombre": "Total patrimonio bruto",
        "descripcion": (
            "Valor total de todos tus bienes a 31-dic-2023: cuentas bancarias, "
            "inversiones, inmuebles, vehiculos, etc. "
            "Verifica con tus extractos al cierre del anio."
        ),
        "seccion": "Patrimonio",
        "requiere_verificacion": True,
    },
    "c30_deudas": {
        "numero": 30,
        "nombre": "Deudas",
        "descripcion": (
            "Saldo total de deudas a 31-dic-2023: credito hipotecario, "
            "tarjetas de credito, prestamos, etc."
        ),
        "seccion": "Patrimonio",
        "requiere_verificacion": True,
    },
    # Cedula General - Rentas de Trabajo
    "c32_ingresos_laborales": {
        "numero": 32,
        "nombre": "Ingresos brutos Rentas de Trabajo",
        "descripcion": (
            "Total salarios, primas, cesantias, bonificaciones y demas pagos "
            "laborales recibidos en 2023. Tomado del Formato 220 de tu(s) empleador(es)."
        ),
        "seccion": "Rentas de Trabajo",
        "requiere_verificacion": False,
    },
    "c33_no_const_laboral": {
        "numero": 33,
        "nombre": "Ingresos no constitutivos de renta - Trabajo",
        "descripcion": (
            "Aportes obligatorios a salud y pension descontados de tu salario. "
            "No son renta gravable."
        ),
        "seccion": "Rentas de Trabajo",
        "requiere_verificacion": False,
    },
    "c35_exenta_afc_fpv": {
        "numero": 35,
        "nombre": "Renta exenta - Aportes AFC/FVP/AVC",
        "descripcion": (
            f"Aportes voluntarios a AFC, Fondos de Pensiones Voluntarias o AVC. "
            f"Exentos hasta el 30%% del ingreso tributario o {3800*UVT/1e6:.1f}M (3.800 UVT)."
        ),
        "seccion": "Rentas de Trabajo",
        "requiere_verificacion": True,
    },
    "c36_otras_exentas_lab": {
        "numero": 36,
        "nombre": "Renta exenta 25% (art. 206 num. 10 ET)",
        "descripcion": (
            f"El 25%% de tus ingresos laborales netos esta exento. "
            f"Tope maximo: $33.505.480 (790 UVT)."
        ),
        "seccion": "Rentas de Trabajo",
        "requiere_verificacion": False,
    },
    "c38_ded_intereses_viv": {
        "numero": 38,
        "nombre": "Deduccion intereses credito hipotecario",
        "descripcion": (
            f"Intereses pagados en 2023 por credito hipotecario de vivienda. "
            f"Maximo deducible: $50.894.400 (1.200 UVT). "
            f"Requiere certificado del banco."
        ),
        "seccion": "Rentas de Trabajo",
        "requiere_verificacion": True,
    },
    "c39_otras_ded_lab": {
        "numero": 39,
        "nombre": "Otras deducciones - Trabajo",
        "descripcion": (
            "Incluye: medicina prepagada (max $8.143.104/ano), "
            "dependientes economicos (72 UVT/persona, max 4), "
            "intereses ICETEX (max $4.241.200), "
            "50%% del GMF certificado por el banco."
        ),
        "seccion": "Rentas de Trabajo",
        "requiere_verificacion": True,
    },
    "c41_exentas_limitadas": {
        "numero": 41,
        "nombre": "Total rentas exentas y deducciones (LIMITADAS)",
        "descripcion": (
            f"Suma de rentas exentas + deducciones, "
            f"limitada al 40%% del ingreso neto o $56.832.080 (1.340 UVT). "
            f"Este limite se aplica automaticamente."
        ),
        "seccion": "Rentas de Trabajo",
        "requiere_verificacion": False,
    },
    "c42_renta_liq_ord_lab": {
        "numero": 42,
        "nombre": "Renta liquida ordinaria - Trabajo",
        "descripcion": (
            "Casilla 34 menos casilla 41. "
            "Base gravable de tus rentas laborales."
        ),
        "seccion": "Rentas de Trabajo",
        "requiere_verificacion": False,
    },
    # Rentas de Capital
    "c58_ing_capital": {
        "numero": 58,
        "nombre": "Ingresos brutos Rentas de Capital",
        "descripcion": (
            "Rendimientos financieros, intereses CDT/ahorros, arrendamientos "
            "recibidos, regalias. Tomado de certificados de entidades financieras."
        ),
        "seccion": "Rentas de Capital",
        "requiere_verificacion": True,
    },
    # Rentas No Laborales
    "c74_ing_no_laborales": {
        "numero": 74,
        "nombre": "Ingresos brutos Rentas No Laborales",
        "descripcion": (
            "Todos los ingresos que no clasifican en las demas cedulas: "
            "venta de activos poseidos menos de 2 anios, indemnizaciones, otros."
        ),
        "seccion": "Rentas No Laborales",
        "requiere_verificacion": True,
    },
    # Pensiones
    "c91_ing_pensiones": {
        "numero": 91,
        "nombre": "Ingresos brutos Cedula de Pensiones",
        "descripcion": (
            "Total pension recibida en 2023. Segun certificado del fondo o Colpensiones."
        ),
        "seccion": "Cedula de Pensiones",
        "requiere_verificacion": True,
    },
    "c95_exenta_pension": {
        "numero": 95,
        "nombre": "Renta exenta pension (25%)",
        "descripcion": (
            f"El 25%% de la pension esta exento. "
            f"Tope: $33.505.480 (790 UVT). "
            f"NOTA: las pensiones menores a 1.000 UVT/mes estan totalmente exentas."
        ),
        "seccion": "Cedula de Pensiones",
        "requiere_verificacion": False,
    },
    # Dividendos
    "c100_dividendos": {
        "numero": 100,
        "nombre": "Ingresos cedula dividendos",
        "descripcion": (
            "Dividendos y participaciones recibidas en 2023. "
            "Verificar con el certificado si son gravados o no gravados."
        ),
        "seccion": "Cedula de Dividendos",
        "requiere_verificacion": True,
    },
    # Impuesto y retenciones
    "c125_impuesto_cargo": {
        "numero": 125,
        "nombre": "Impuesto sobre la renta (estimado)",
        "descripcion": (
            "Calculado segun tabla del art. 241 ET sobre la renta liquida gravable. "
            "VALOR ESTIMADO — puede variar segun otros campos que no esten en la exogena."
        ),
        "seccion": "Liquidacion",
        "requiere_verificacion": True,
    },
    "total_retenciones": {
        "numero": None,
        "nombre": "Total retenciones en la fuente",
        "descripcion": (
            "Suma de todas las retenciones que te practicaron durante el anio. "
            "Si supera el impuesto, tienes SALDO A FAVOR."
        ),
        "seccion": "Liquidacion",
        "requiere_verificacion": False,
    },
    "saldo_cargo_o_favor": {
        "numero": None,
        "nombre": "Saldo a cargo (+) o a favor (-)",
        "descripcion": (
            "Positivo = debes pagar. Negativo = tienes saldo a favor (DIAN te devuelve). "
            "VALOR ESTIMADO — verifica en el sistema de diligenciamiento de la DIAN."
        ),
        "seccion": "Liquidacion",
        "requiere_verificacion": True,
    },
}


class FormGenerator210(IFormGenerator):
    """
    Genera el borrador del Formulario 210 en Excel.
    Dos hojas: resumen ejecutivo y detalle por casilla.
    """

    def generar_excel(self, borrador: dict, ruta_salida: str) -> str:
        try:
            import openpyxl
            from openpyxl.styles import (
                PatternFill, Font, Alignment, Border, Side, numbers
            )
            from openpyxl.utils import get_column_letter
        except ImportError:
            raise RuntimeError("Instala openpyxl: pip install openpyxl")

        wb = openpyxl.Workbook()

        # --- Hoja 1: Resumen ejecutivo ---
        ws_resumen = wb.active
        ws_resumen.title = "Resumen"
        self._llenar_hoja_resumen(ws_resumen, borrador)

        # --- Hoja 2: Detalle por casilla ---
        ws_detalle = wb.create_sheet("Detalle Casillas")
        self._llenar_hoja_detalle(ws_detalle, borrador)

        wb.save(ruta_salida)
        logger.info(f"Excel generado en: {ruta_salida}")
        return ruta_salida

    def _llenar_hoja_resumen(self, ws, borrador: dict):
        """Hoja ejecutiva con los totales mas importantes."""
        from openpyxl.styles import PatternFill, Font, Alignment

        # Titulo
        ws["A1"] = "BORRADOR FORMULARIO 210 - AÑO GRAVABLE {ANNO_GRAVABLE}"
        ws["A1"].font = Font(bold=True, size=14)
        ws["A1"].fill = PatternFill("solid", fgColor="1F4E79")
        ws["A1"].font = Font(bold=True, size=14, color="FFFFFF")
        ws.merge_cells("A1:D1")

        ws["A2"] = f"UVT 2023: ${UVT:,}"
        ws["A2"].font = Font(italic=True, size=10)
        ws["A3"] = "IMPORTANTE: Este es un borrador estimativo. Verifica en el sistema DIAN."
        ws["A3"].font = Font(italic=True, color="FF0000")

        encabezados = ["Seccion", "Campo", "Valor estimado ($)", "Estado"]
        for col, enc in enumerate(encabezados, 1):
            celda = ws.cell(row=5, column=col, value=enc)
            celda.font = Font(bold=True, color="FFFFFF")
            celda.fill = PatternFill("solid", fgColor="2E75B6")

        fila = 6
        secciones_mostrar = [
            ("Patrimonio",        "c29_patrimonio_bruto",    "Patrimonio bruto"),
            ("Rentas de Trabajo", "c32_ingresos_laborales",  "Ingresos laborales"),
            ("Rentas de Trabajo", "c42_renta_liq_ord_lab",   "Renta liquida laboral"),
            ("Rentas de Capital", "c58_ing_capital",         "Ingresos capital"),
            ("Rentas No Lab.",    "c74_ing_no_laborales",    "Ingresos no laborales"),
            ("Pensiones",         "c91_ing_pensiones",       "Ingresos pensiones"),
            ("Dividendos",        "c100_dividendos",         "Dividendos"),
            ("Liquidacion",       "renta_liq_cedula_general","Renta liquida cedula general"),
            ("Liquidacion",       "c125_impuesto_cargo",     "Impuesto estimado"),
            ("Liquidacion",       "total_retenciones",       "Total retenciones"),
            ("Liquidacion",       "saldo_cargo_o_favor",     "Saldo a cargo/favor"),
        ]

        for seccion, campo, etiqueta in secciones_mostrar:
            valor = borrador.get(campo, 0)
            if not valor:
                continue

            ws.cell(row=fila, column=1, value=seccion)
            ws.cell(row=fila, column=2, value=etiqueta)
            celda_valor = ws.cell(row=fila, column=3, value=valor)
            celda_valor.number_format = "#,##0"

            if campo == "saldo_cargo_o_favor":
                if isinstance(valor, (int, float)):
                    if valor > 0:
                        ws.cell(row=fila, column=4, value="A CARGO (debe pagar)")
                        celda_valor.fill = PatternFill("solid", fgColor="FFE0E0")
                    elif valor < 0:
                        ws.cell(row=fila, column=4, value="A FAVOR (devolucion)")
                        celda_valor.fill = PatternFill("solid", fgColor="E0FFE0")
                    else:
                        ws.cell(row=fila, column=4, value="Saldo cero")
            else:
                info = _CASILLAS_210.get(campo, {})
                estado = "Requiere verificar" if info.get("requiere_verificacion") else "Calculado"
                ws.cell(row=fila, column=4, value=estado)

            fila += 1

        # Ajustar ancho de columnas
        ws.column_dimensions["A"].width = 20
        ws.column_dimensions["B"].width = 35
        ws.column_dimensions["C"].width = 22
        ws.column_dimensions["D"].width = 25

    def _llenar_hoja_detalle(self, ws, borrador: dict):
        """Hoja detallada con cada casilla del 210."""
        from openpyxl.styles import PatternFill, Font, Alignment, numbers

        encabezados = [
            "Casilla #", "Nombre del Campo", "Valor ($)", "Seccion",
            "Explicacion", "Requiere verificacion?"
        ]
        for col, enc in enumerate(encabezados, 1):
            celda = ws.cell(row=1, column=col, value=enc)
            celda.font = Font(bold=True, color="FFFFFF")
            celda.fill = PatternFill("solid", fgColor="1F4E79")

        fila = 2
        secciones_vistas = set()

        for campo_key, info in _CASILLAS_210.items():
            valor = borrador.get(campo_key)
            if valor is None:
                continue

            seccion = info.get("seccion", "")
            # Fila separadora por seccion
            if seccion not in secciones_vistas:
                ws.cell(row=fila, column=1, value=seccion)
                ws.merge_cells(f"A{fila}:F{fila}")
                ws.cell(row=fila, column=1).fill = PatternFill("solid", fgColor="D6E4F0")
                ws.cell(row=fila, column=1).font = Font(bold=True)
                secciones_vistas.add(seccion)
                fila += 1

            num  = info.get("numero")
            nombre = info.get("nombre", campo_key)
            desc   = info.get("descripcion", "")
            verif  = "SI" if info.get("requiere_verificacion") else "No"

            ws.cell(row=fila, column=1, value=num or "")
            ws.cell(row=fila, column=2, value=nombre)
            celda_valor = ws.cell(row=fila, column=3, value=valor if isinstance(valor, (int, float)) else str(valor))
            if isinstance(valor, (int, float)):
                celda_valor.number_format = "#,##0"
                if campo_key == "saldo_cargo_o_favor" and valor < 0:
                    celda_valor.fill = PatternFill("solid", fgColor="C6EFCE")
                elif campo_key == "saldo_cargo_o_favor" and valor > 0:
                    celda_valor.fill = PatternFill("solid", fgColor="FFC7CE")

            ws.cell(row=fila, column=4, value=seccion)
            celda_desc = ws.cell(row=fila, column=5, value=desc)
            celda_desc.alignment = Alignment(wrap_text=True)

            celda_verif = ws.cell(row=fila, column=6, value=verif)
            if verif == "SI":
                celda_verif.fill = PatternFill("solid", fgColor="FFF2CC")

            ws.row_dimensions[fila].height = 45
            fila += 1

        # Ajustar columnas
        ws.column_dimensions["A"].width = 10
        ws.column_dimensions["B"].width = 40
        ws.column_dimensions["C"].width = 20
        ws.column_dimensions["D"].width = 22
        ws.column_dimensions["E"].width = 70
        ws.column_dimensions["F"].width = 22

    def generar_resumen_pdf(self, borrador: dict, ruta_salida: str) -> str:
        """Genera PDF de resumen (implementacion basica)."""
        try:
            from reportlab.lib.pagesizes import letter
            from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer
            from reportlab.lib.styles import getSampleStyleSheet

            doc    = SimpleDocTemplate(ruta_salida, pagesize=letter)
            styles = getSampleStyleSheet()
            story  = []

            story.append(Paragraph(
                "Borrador Formulario 210 - Año Gravable 2023",
                styles["Title"]
            ))
            story.append(Spacer(1, 12))

            saldo = borrador.get("saldo_cargo_o_favor", 0)
            if isinstance(saldo, (int, float)):
                estado = "SALDO A FAVOR" if saldo < 0 else "SALDO A CARGO"
                color  = "green" if saldo < 0 else "red"
                story.append(Paragraph(
                    f'<font color="{color}"><b>{estado}: ${abs(saldo):,.0f}</b></font>',
                    styles["Heading2"]
                ))

            story.append(Spacer(1, 12))
            story.append(Paragraph(
                "Consulta el archivo Excel para el detalle completo "
                "de cada casilla con su explicacion.",
                styles["Normal"]
            ))

            doc.build(story)
            return ruta_salida
        except ImportError:
            logger.warning("reportlab no disponible, omitiendo PDF")
            return ""
