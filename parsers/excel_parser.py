"""
parsers/excel_parser.py
Lee el reporte de exogena descargado del portal DIAN.

El archivo tiene dos posibles formatos:

  FORMATO A (portal moderno — reporte consolidado):
    Hoja unica con cabecera de metadatos y columnas:
    NIT | Nombre | NIT | Nombre reportado | Detalle | Valor |
    Uso declaracion Sugerida | Informacion Adicional
    La columna 'Uso declaracion Sugerida' contiene codigos R{n} que
    mapean directamente a casillas del Formulario 210.

  FORMATO B (formato clasico con columnas concepto/formato):
    Hojas por formato (220, 1001, etc.) con columnas como
    concepto, valor, retencion, nit_informante, etc.

El parser detecta automaticamente cual formato tiene el archivo.
"""
import re
import logging
from typing import Optional
import pandas as pd
from interfaces.base import IExogenaParser, ResumenExogena
from config.constants import (
    ENTIDADES_FINANCIERAS,
    ENTIDADES_PENSION,
    ENTIDADES_FPV,
    ENTIDADES_AFC,
    FORMATOS_EXOGENA,
)

# Alias para compatibilidad con imports anteriores
ENTIDADES_FINANCIERAS_CONOCIDAS = ENTIDADES_FINANCIERAS

logger = logging.getLogger(__name__)

# ─────────────────────────────────────────────────────────────
# FORMATO A — mapeo casilla → campo del ResumenExogena
# Se usa el codigo R{n} extraido de 'Uso declaracion Sugerida'
# Los codigos vienen del ET y son estables entre años
# ─────────────────────────────────────────────────────────────
_CASILLA_A_CAMPO = {
    32:  "ingresos_laborales",        # Ingresos brutos rentas de trabajo
    43:  "ingresos_no_laborales_trabajo",  # Trabajo no laboral (honorarios con costos)
    58:  "ingresos_capital",          # Ingresos brutos rentas de capital
    74:  "ingresos_no_laborales",     # Rentas no laborales
    99:  "ingresos_pensiones",        # Ingresos brutos pensiones
    33:  "ingresos_no_const",         # Ingresos no constitutivos trabajo
    44:  "ingresos_no_const",         # Ingresos no constitutivos trabajo no laboral
    59:  "ingresos_no_const",         # Ingresos no constitutivos capital
    100: "ingresos_no_const",         # Ingresos no constitutivos pensiones
    35:  "rentas_exentas_trabajo",    # Aportes AFC/FPV rentas trabajo
    36:  "rentas_exentas_trabajo",    # Otras rentas exentas trabajo
    47:  "rentas_exentas_trabajo",    # Aportes AFC/FPV trabajo no laboral
    63:  "rentas_exentas_capital",    # Aportes AFC/FPV capital
    132: "retenciones_trabajo",       # Retenciones año gravable
    29:  "patrimonio_bruto",          # Patrimonio bruto
    28:  "factura_electronica",       # Factura electronica 1%
}

# Filas de resumen/topes al final del archivo — se ignoran
# La DIAN las agrega como totales orientativos, no son datos reales
_PATRONES_IGNORAR_USO = [
    r"tope\s+[1-5]",          # Tope 1 - Ingresos, Tope 2 - Patrimonio, etc.
    r"tope\s+[1-5]\s*[:\-\.]", # Variantes con separador
]

# Cuando no hay codigo R, mapear por texto del campo Uso
# Solo para casos que la DIAN nunca pone con R{n}
_TEXTO_USO_A_CAMPO = {
    r"consignaciones\s+e\s+inversiones":  "total_consignaciones",
    r"consumos\s+tc":                     "total_compras",
    r"ingresos\s+no\s+constitutivos":     "ingresos_no_const",
    r"rentas\s+exentas.*rentas\s+":       "rentas_exentas_trabajo",
}

# ─────────────────────────────────────────────────────────────
# FORMATO B — aliases de columnas (formato clasico)
# ─────────────────────────────────────────────────────────────
_ALIAS_COLUMNAS = {
    "nit_informante":    ["nit", "nit informante", "nit_informante", "identificacion"],
    "nombre_informante": ["nombre", "razon social", "nombre informante", "razon_social"],
    "concepto":          ["concepto", "codigo concepto", "cod concepto"],
    "valor":             ["valor", "monto", "valor pagado", "valor bruto"],
    "valor_retencion":   ["retencion", "valor retencion", "retenido", "valor_retencion"],
    "formato":           ["formato", "tipo reporte", "formato reporte"],
    "nit_receptor":      ["nit receptor", "nit beneficiario", "cedula"],
    "nombre_receptor":   ["nombre receptor", "nombre beneficiario"],
}

_CONCEPTOS_TRABAJO         = {"5001","5002","5003","5004","5005","5009","5011","5012","5013","22","23","24"}
_CONCEPTOS_INDEPENDIENTE   = {"1001","1002","1003","1004","1005","1006","1007","1008"}
_CONCEPTOS_RENDIMIENTOS    = {"1404","1405","1406","1403"}
_CONCEPTOS_PENSION         = {"5004","5051","5052"}
_CONCEPTOS_NO_CONSTITUTIVOS= {"2204","2205","2206","2207"}
_CONCEPTOS_GMF             = {"1115"}


class ExogenaParser(IExogenaParser):

    def parsear(self, ruta_archivo: str) -> ResumenExogena:
        logger.info(f"Parseando exogena: {ruta_archivo}")
        resultado = ResumenExogena()

        try:
            xls = pd.ExcelFile(ruta_archivo)
        except Exception as e:
            raise ValueError(f"No se pudo abrir el archivo Excel: {e}")

        # Detectar formato leyendo la primera hoja con openpyxl (sin pandas)
        if self._es_formato_portal_moderno(ruta_archivo):
            logger.info("Formato detectado: portal DIAN moderno (Uso declaracion Sugerida)")
            self._parsear_formato_moderno(ruta_archivo, resultado)
        else:
            logger.info("Formato detectado: exogena clasica (columnas concepto/formato)")
            hojas_procesadas = 0
            for hoja in xls.sheet_names:
                try:
                    df = pd.read_excel(ruta_archivo, sheet_name=hoja, dtype=str)
                    df.columns = [str(c).strip().lower() for c in df.columns]
                    df = df.fillna("0")
                    if self._es_hoja_relevante(df):
                        self._procesar_hoja_clasica(df, hoja, resultado)
                        hojas_procesadas += 1
                except Exception as e:
                    logger.warning(f"Error procesando hoja '{hoja}': {e}")
                    resultado.advertencias.append(f"Hoja '{hoja}' no pudo procesarse: {e}")

            if hojas_procesadas == 0:
                raise ValueError(
                    "El archivo no tiene hojas con datos reconocibles de exogena DIAN. "
                    "Verifica que sea el archivo correcto descargado del portal."
                )

        self._calcular_totales(resultado)
        logger.info(
            f"Exogena procesada: ingresos_lab={resultado.ingresos_laborales:,.0f} "
            f"capital={resultado.ingresos_capital:,.0f} "
            f"retenciones={resultado.total_retenciones:,.0f} "
            f"pagadores={len(resultado.pagadores)}"
        )
        return resultado

    # ══════════════════════════════════════════════════════════
    # FORMATO A — portal DIAN moderno
    # ══════════════════════════════════════════════════════════

    def _es_formato_portal_moderno(self, ruta: str) -> bool:
        """
        Detecta si el archivo tiene el formato del portal moderno
        buscando la columna 'Uso declaracion Sugerida' o 'Detalle'.
        No asume posicion de fila — busca en todas.
        """
        try:
            import openpyxl
            wb = openpyxl.load_workbook(ruta, read_only=True)
            ws = wb[wb.sheetnames[0]]
            for row in ws.iter_rows(max_row=30, values_only=True):
                fila = " ".join(str(v).lower() for v in row if v)
                if "uso declaraci" in fila or "detalle" in fila:
                    return True
        except Exception:
            pass
        return False

    def _parsear_formato_moderno(self, ruta: str, resultado: ResumenExogena) -> None:
        import openpyxl
        wb = openpyxl.load_workbook(ruta, read_only=True, data_only=True)
        ws = wb[wb.sheetnames[0]]
        rows = list(ws.iter_rows(values_only=True))

        # Encontrar fila de headers: la que tiene mas celdas no nulas
        header_idx = max(
            range(len(rows)),
            key=lambda i: sum(1 for v in rows[i] if v is not None)
        )
        headers = [str(v).strip().lower() if v else "" for v in rows[header_idx]]

        # Ubicar columnas por nombre, no por indice fijo
        col = {
            "nit_pagador":  self._idx_col(headers, ["nit"]),
            "nom_pagador":  self._idx_col(headers, ["nombre / razón social", "nombre / razon social", "nombre"]),
            "detalle":      self._idx_col(headers, ["detalle"]),
            "valor":        self._idx_col(headers, ["valor"]),
            "uso":          self._idx_col(headers, ["uso declaración sugerida", "uso declaracion sugerida", "uso"]),
        }

        # Extraer metadatos del declarante de las filas anteriores al header
        nit, nombre, anno = self._extraer_metadatos_moderno(rows[:header_idx])
        resultado.nit_usuario    = nit
        resultado.nombre_usuario = nombre

        pagadores  = {}
        entidades  = {}

        for row in rows[header_idx + 1:]:
            if not row or all(v is None for v in row):
                continue

            def cel(idx):
                return row[idx] if idx is not None and idx < len(row) else None

            nit_pag  = str(cel(col["nit_pagador"]) or "").strip()
            nom_pag  = str(cel(col["nom_pagador"])  or "").strip()
            detalle  = str(cel(col["detalle"])      or "").strip()
            valor    = self._a_float(cel(col["valor"]))
            uso      = str(cel(col["uso"])          or "").strip()

            if valor == 0.0:
                continue

            # Ignorar filas de resumen/topes (no son datos reales)
            if self._es_fila_resumen(uso, detalle):
                continue

            # Determinar campo destino a partir del codigo R{n} en 'Uso'
            campo = self._campo_desde_uso(uso)

            if campo:
                self._acumular(resultado, campo, valor)

                # Registrar pagadores laborales
                if campo == "ingresos_laborales" and nit_pag and valor > 0:
                    if nit_pag not in pagadores:
                        pagadores[nit_pag] = {"nombre": nom_pag, "valor": 0.0, "tipo": "trabajo_laboral"}
                    pagadores[nit_pag]["valor"] += valor

                # Registrar entidades financieras
                if campo in ("ingresos_capital", "retenciones_trabajo") and nit_pag:
                    entidades[nit_pag] = nom_pag

        resultado.pagadores = [
            {"nombre": v["nombre"], "valor": v["valor"], "tipo": v["tipo"]}
            for v in pagadores.values() if v["valor"] > 0
        ]
        resultado.entidades_financieras = [
            {"nombre": nom} for nom in entidades.values()
        ]

    def _campo_desde_uso(self, uso: str) -> Optional[str]:
        """
        Extrae el campo del ResumenExogena leyendo el codigo R{n}
        de la columna 'Uso declaracion Sugerida'.

        Estrategia:
        1. Buscar patron R{numero} — es lo estable del ET.
           Si lo encuentra, ese es el campo. Los topes ('Tope 1', etc.)
           que acompanan el codigo son solo etiquetas orientativas
           y no impiden usar el codigo R.
        2. Si no hay codigo R, verificar si es un tope puro → ignorar.
        3. Si no hay codigo R pero hay texto conocido → mapear por texto.
        4. Sin match → None (fila se ignora).
        """
        # Extraer primer codigo R{n} — tiene prioridad sobre todo
        match = re.search(r'\bR(\d{2,3})\b', uso, re.IGNORECASE)
        if match:
            casilla = int(match.group(1))
            campo = _CASILLA_A_CAMPO.get(casilla)
            if campo:
                return campo
            logger.debug(f"Casilla R{casilla} sin mapeo en uso: '{uso}'")
            return None

        # Sin codigo R — ignorar topes puros
        uso_lower = uso.lower()
        for patron in _PATRONES_IGNORAR_USO:
            if re.search(patron, uso_lower):
                return None

        # Sin codigo R y no es tope — intentar por texto
        for patron, campo in _TEXTO_USO_A_CAMPO.items():
            if re.search(patron, uso_lower):
                return campo

        return None

    def _es_fila_resumen(self, uso: str, detalle: str) -> bool:
        """
        Detecta filas de resumen/totales que la DIAN agrega al final
        del archivo y que NO son datos reales a declarar.

        IMPORTANTE: filas con 'Tope 1' pero que tambien tienen R32, R58 etc.
        NO son resumenes — son datos reales con etiqueta de tope.
        Solo se ignoran las filas de tope SIN codigo R (los resumenes puros).
        """
        uso_lower     = uso.lower()
        detalle_lower = detalle.lower()

        # Si tiene codigo R → es dato real, nunca ignorar
        if re.search(r'\bR\d{2,3}\b', uso, re.IGNORECASE):
            return False

        # Tope sin codigo R → resumen orientativo, ignorar
        for patron in _PATRONES_IGNORAR_USO:
            if re.search(patron, uso_lower):
                return True

        # Filas de totales al final sin codigo R
        resumen_detalle = {
            "total patrimonio bruto declarado",
            "monto total de facturación electrónica",
            "monto total de facturaci",
            "suma valor total facturas",
            "total adquisiciones consumos",
        }
        for r in resumen_detalle:
            if r in detalle_lower:
                return True

        return False

    def _extraer_metadatos_moderno(self, rows: list) -> tuple:
        """Extrae NIT, nombre y año de las filas de cabecera del formato moderno."""
        nit    = ""
        nombre = ""
        anno   = 2025

        for row in rows:
            if not row:
                continue
            fila = " ".join(str(v) for v in row if v is not None)
            fila_l = fila.lower()

            if "identificación:" in fila_l or "identificacion:" in fila_l:
                for v in row:
                    s = str(v).strip() if v else ""
                    if s.isdigit() and len(s) >= 6:
                        nit = s
                        break

            if "nombres" in fila_l and "razón" in fila_l:
                for v in row:
                    s = str(v).strip() if v else ""
                    if (s and len(s) > 5
                            and not any(k in s.lower() for k in
                                        ["nombres", "razón", "razon", "social", ":"])):
                        nombre = s
                        break

            if "año al que" in fila_l or "ano al que" in fila_l:
                for v in row:
                    s = str(v).strip() if v else ""
                    if s.isdigit() and 2020 <= int(s) <= 2030:
                        anno = int(s)
                        break

        return nit, nombre, anno

    @staticmethod
    def _idx_col(headers: list, candidatos: list) -> Optional[int]:
        """Encuentra el indice de una columna buscando por nombre, sin asumir posicion."""
        for candidato in candidatos:
            for i, h in enumerate(headers):
                if candidato in h:
                    return i
        return None

    @staticmethod
    def _acumular(resultado: ResumenExogena, campo: str, valor: float) -> None:
        actual = getattr(resultado, campo, 0.0) or 0.0
        setattr(resultado, campo, actual + valor)

    # ══════════════════════════════════════════════════════════
    # FORMATO B — exogena clasica (concepto/formato)
    # ══════════════════════════════════════════════════════════

    def _es_hoja_relevante(self, df: pd.DataFrame) -> bool:
        columnas = set(df.columns)
        tiene_valor = any(any(a in c for a in ["valor", "monto"]) for c in columnas)
        tiene_id    = any(any(a in c for a in ["nit", "cedula", "identificacion"]) for c in columnas)
        return tiene_valor and tiene_id and len(df) > 0

    def _normalizar_col(self, df: pd.DataFrame, campo: str) -> Optional[str]:
        aliases = _ALIAS_COLUMNAS.get(campo, [])
        for alias in aliases:
            for col in df.columns:
                if alias in col:
                    return col
        return None

    def _procesar_hoja_clasica(self, df: pd.DataFrame, nombre_hoja: str,
                                resultado: ResumenExogena) -> None:
        col_nombre   = self._normalizar_col(df, "nombre_informante")
        col_nit      = self._normalizar_col(df, "nit_informante")
        col_valor    = self._normalizar_col(df, "valor")
        col_reten    = self._normalizar_col(df, "valor_retencion")
        col_concepto = self._normalizar_col(df, "concepto")
        col_formato  = self._normalizar_col(df, "formato")

        if not col_valor:
            return

        for _, fila in df.iterrows():
            nombre    = str(fila.get(col_nombre, "")).strip().upper() if col_nombre else ""
            nit       = str(fila.get(col_nit, "")).strip()            if col_nit    else ""
            valor     = self._limpiar_valor(fila.get(col_valor, "0"))
            retencion = self._limpiar_valor(fila.get(col_reten, "0")) if col_reten else 0.0
            concepto  = str(fila.get(col_concepto, "")).strip()        if col_concepto else ""
            formato   = str(fila.get(col_formato, "")).strip()         if col_formato  else ""

            if valor == 0.0:
                continue

            if not resultado.nit_usuario:
                col_nit_rec = self._normalizar_col(df, "nit_receptor")
                if col_nit_rec:
                    resultado.nit_usuario = str(fila.get(col_nit_rec, "")).strip()

            if not resultado.nombre_usuario:
                col_nom_rec = self._normalizar_col(df, "nombre_receptor")
                if col_nom_rec:
                    resultado.nombre_usuario = str(fila.get(col_nom_rec, "")).strip().upper()

            tipo = self._clasificar_tipo_clasico(nombre, concepto, formato, nombre_hoja)

            mapa = {
                "trabajo_laboral":       ("ingresos_laborales",             "retenciones_trabajo"),
                "trabajo_no_laboral":    ("ingresos_no_laborales_trabajo",  "retenciones_no_laborales"),
                "capital":               ("ingresos_capital",               "retenciones_capital"),
                "no_laboral":            ("ingresos_no_laborales",          "retenciones_no_laborales"),
                "pension":               ("ingresos_pensiones",             "retenciones_pensiones"),
                "dividendo":             ("dividendos",                     None),
                "ganancia_ocasional":    ("ganancias_ocasionales",          None),
                "patrimonio_cuenta":     ("saldos_cuentas",                 None),
                "no_constitutivo":       ("ingresos_no_const",              None),
                "gmf":                   ("gmf_pagado",                     None),
                "consignacion":          ("total_consignaciones",           None),
            }

            if tipo in mapa:
                campo_ingreso, campo_ret = mapa[tipo]
                self._acumular(resultado, campo_ingreso, valor)
                if campo_ret:
                    self._acumular(resultado, campo_ret, retencion)

                if tipo == "dividendo":
                    resultado.tiene_dividendos = True

                if nombre and valor > 0 and tipo not in (
                    "patrimonio_cuenta", "gmf", "consignacion", "no_constitutivo"
                ):
                    self._registrar_pagador(resultado, nombre, nit, valor, retencion, tipo)

    def _clasificar_tipo_clasico(self, nombre: str, concepto: str,
                                  formato: str, hoja: str) -> str:
        hoja_l      = hoja.lower()
        formato_num = re.sub(r"[^\d]", "", formato)

        if formato_num == "220" or "220" in hoja_l:          return "trabajo_laboral"
        if formato_num in ("2276","1001"):                    return "trabajo_no_laboral"
        if formato_num == "2278":                             return "capital"
        if formato_num == "1012":                             return "patrimonio_cuenta"
        if formato_num == "1010":                             return "dividendo"
        if "1115" in concepto or "gmf" in hoja_l:            return "gmf"
        if "consignaci" in hoja_l:                            return "consignacion"
        if any(e in nombre for e in ENTIDADES_PENSION):       return "pension"
        if concepto in _CONCEPTOS_PENSION:                    return "pension"
        if concepto in _CONCEPTOS_TRABAJO:                    return "trabajo_laboral"
        if concepto in _CONCEPTOS_INDEPENDIENTE:              return "trabajo_no_laboral"
        if concepto in _CONCEPTOS_RENDIMIENTOS:               return "capital"
        if concepto in _CONCEPTOS_NO_CONSTITUTIVOS:           return "no_constitutivo"
        if concepto in _CONCEPTOS_GMF:                        return "gmf"
        if any(e in nombre for e in ENTIDADES_FINANCIERAS):   return "capital"
        return "no_laboral"

    def _registrar_pagador(self, resultado, nombre, nit, valor, retencion, tipo):
        clave = nit or nombre
        for p in resultado.pagadores:
            if p.get("clave") == clave:
                p["valor"]    += valor
                p["retencion"] = p.get("retencion", 0) + retencion
                return
        resultado.pagadores.append({
            "clave": clave, "nombre": nombre, "nit": nit,
            "valor": valor, "retencion": retencion, "tipo": tipo,
        })
        if any(e in nombre for e in ENTIDADES_FINANCIERAS):
            if not any(e["nombre"] == nombre for e in resultado.entidades_financieras):
                resultado.entidades_financieras.append({"nombre": nombre, "nit": nit})

    # ══════════════════════════════════════════════════════════
    # Utilidades comunes
    # ══════════════════════════════════════════════════════════

    def _calcular_totales(self, resultado: ResumenExogena) -> None:
        resultado.total_ingresos_brutos = (
            resultado.ingresos_laborales
            + resultado.ingresos_no_laborales_trabajo
            + resultado.ingresos_capital
            + resultado.ingresos_no_laborales
            + resultado.ingresos_pensiones
            + resultado.dividendos
        )
        resultado.total_retenciones = (
            resultado.retenciones_trabajo
            + resultado.retenciones_capital
            + resultado.retenciones_no_laborales
            + resultado.retenciones_pensiones
        )

    @staticmethod
    def _limpiar_valor(v) -> float:
        if not v or str(v) in ("0", "", "nan", "none", "None"):
            return 0.0
        s = re.sub(r"[^\d\-]", "", str(v).replace(",", "").replace("$", ""))
        try:
            return float(s)
        except (ValueError, TypeError):
            return 0.0

    @staticmethod
    def _a_float(v) -> float:
        if v is None:
            return 0.0
        if isinstance(v, (int, float)):
            return float(v)
        s = re.sub(r"[^\d\-\.]", "", str(v).replace(",", "."))
        try:
            return float(s)
        except ValueError:
            return 0.0